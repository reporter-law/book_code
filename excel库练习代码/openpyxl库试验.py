# -*-  coding: utf-8 -*-
# Author: xxx
# Datetime : 2019/12/12 10:25
# software: PyCharm
# usage:
import openpyxl
from tqdm import tqdm

wb = openpyxl.Workbook()
ws = wb.active
ws['A1']=12552

ws.append([222])#在下一行加入,必须是列表
for i in tqdm(range(100)):
    ws.append([i])
ws.cell(row=4, column=2, value=10)
for i in range(1,10):
    for w in range(1,10):
        ws.cell(row=i,column=w,value=i)

wb.save('uuyyuy.xlsx')#这个表不可以打开状态下添加