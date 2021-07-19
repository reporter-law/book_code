# -*-  coding: utf-8 -*-
# Author: xxx
# Datetime : 2019/12/12 10:25
# software: PyCharm
# usage:
import requests
from bs4 import BeautifulSoup
from tqdm import tqdm
import lxml
from openpyxl import Workbook, load_workbook
import openpyxl



#url分析：
'''url1 = http://wenshu.court.gov.cn/website/wenshu/181217BMTKHNT2W0/index.html?pageId=6d451335753c8fce6c4400903e4ad6b8&s8=02

url2 = http://wenshu.court.gov.cn/website/wenshu/181217BMTKHNT2W0/index.html?pageId=6d451335753c8fce6c4400903e4ad6b8&s8=02

url3 = http://wenshu.court.gov.cn/website/wenshu/181217BMTKHNT2W0/index.html?pageId=6d451335753c8fce6c4400903e4ad6b8&s8=02

url4 = http://wenshu.court.gov.cn/website/wenshu/181217BMTKHNT2W0/index.html?pageId=6d451335753c8fce6c4400903e4ad6b8&s8=02
首页：http://wenshu.court.gov.cn/website/wenshu/181029CR4M5A62CH/index.html
url 无法分析'''

#url = 'http://wenshu.court.gov.cn/website/wenshu/181217BMTKHNT2W0/index.html?pageId=6d451335753c8fce6c4400903e4ad6b8&s8=02'

#嵩天老师教程：寻找其他路径
#天同url = 'https://www.jufaanli.com/search2?TypeKey=1:%E5%88%91%E4%BA%8B&Page=2'也找不到
#最高法：url = 'http://www.court.gov.cn/wenshu.html'可能不全
'''法律图书馆：url ='http://www.law-lib.com/cpws/cpwsml-cx.asp'
步骤：第一步：内容多网页查找，寻找一个或者几个url的组合；第二步：网页分析；第三步：网页内容存储'''
#url = url ='http://www.law-lib.com/cpws/cpwsml-cx.asp'
def get_HTMLTEXT(url):
    try:
        headers = {'User-Agent' : 'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'}
        r = requests.get(url,headers=headers,timeout = 30)
        r.raise_for_status()
        r.encoding = r.apparent_encoding#程序判断encoding的编码方式,出现解码错误，原因直接使用utf-8
        return r.text
    except:
        print('爬取失败')
        print(r.status_code)

#print(get_text())
#问题需要javescript

#信息存储库
def getlist(html):
    #count = 1
    soup = BeautifulSoup(html, 'lxml')
    criminal_jud_lists = []
    #print('hello')
    #try:
    for contents in soup.find_all('ul',attrs={'class':'line2'}):
        for content in contents.find_all('a', attrs={'target':'_blank'}):
            content_1 = content.text
        #print(content)
    #for content in contents.a.text:
    #print(len(contents_soup))#长度只有1？_-----contents_soup = soup.findAll('ul', attrs = {"class":"line2" })
        #print(len(content))

        #contents = contents_soup.findAll('a', attrs={'target':'_blank'}).text
        #只找到前面的ul.li.a而不是后面的
        #print(content)
            criminal_jud_lists.append(content_1)
    #print(criminal_jud_lists)
   # print(criminal_jud_lists)
    #print(criminal_jud_lists)
    return criminal_jud_lists
    #for content in contents:#存在的问题attrs不是直接找到内容只是找到节点
        #attrs={},attrs{]错误
        #findAll返回的是列表形式，不能直接
        #criminal_jud_list = content.text
        #print(criminal_jud_list)#都是空集？又好了！
        #text 与string效果一样，只是可能在不够精确时，text更好


#def get_excel(jud_list):

        #ws.append(jud_list)#必须是list,tuple,range or generator, or a dict,不能是str
        #ws.cell(row = i+1,column =1,value= jud_list[i])
       # ws.append(i,jud_list)#指定行列是否可行,不可行！
                #print(value)
        #count = count + 1
        #print('\r当前速度： {:.2f}%'.format(count * 100 / len(criminal_jud_list)), end='')  # 少了括号
        #wb.save('sample.xlsx')

    #问题在于会覆盖原内容?只能数据汇总后进行吗？
        #print(criminal_jud_list)
            #print(content)
    #except:
        #count = count + 1
        #print('\r当前速度： {:.2f}%'.format(count * 100 /len(criminal_jud_list)), end='')
        #print('失败')
        #continue
    #问题可能有许多a标签？----节点定位
    #出现问题，只写入了一个文书？是否因为只在一个循环中打转！或许由于循环即使被隔开也还是一个循环？

jud_lists =[]
def main(page):
    url =' http://www.law-lib.com/cpws/cpwsml-cx.asp?bbdw=&pages='+ str(number) + '&tm1=&tm2='#嵩天老师这里有两个网页网址，因为从两个网页共同获取
    html = get_HTMLTEXT(url)
    jud_list = getlist(html)
    #print(jud_list)
    jud_lists.append(','.join(jud_list))
    return jud_lists



if __name__== '__main__':
    for number in tqdm(range(1,50)):
        main(number)
    #print(jud_lists)
    wb = openpyxl.Workbook()
    ws = wb.active
    for i in range(len(jud_lists)):
        #print(i),i只49，因为49页
        #print(i)
        #for value in values:
        #print(values[x])
        #print(value)
        ws.cell(row = i+1,column = 1, value = jud_lists[i])
    wb.save('测试文件.xlsx')
#暂时无法完成
#如果要不对values进行遍历，而适用values[]这样，[x]中的x需要循环49次！！！！！0
#for for 循环即使不输出但是循环次数仍在

#暂时仍无法完成

