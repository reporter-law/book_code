import re
import requests
from tqdm import tqdm
import time
import json
import openpyxl
#url解析
#url分析 换页使得?offset=0的数字增加10
'''
def url_list():
    url_lists = []
    for number in tqdm(range(1)):
        url = 'https://maoyan.com/board/4?offset=' + str(number*10)
        url_lists.append(url)
    return url_lists'''
#不知道搞什么一下行一下不行

#print(url_list())

#网页返回
def page(url):
    #texts =[]
    #for url in url_list():
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36'}
    r = requests.get(url,headers = headers)
        #texts.append(r.text)
    return r.text

#print(page())
#内容匹配
def content_one(html):
    content_list = []
    content_lists = re.findall('<dd>.*?board-index.*?>(.*?)</i>'
                               '.*?<img data-src=\"(.*?)\"'
                               '.*?<p class="name">.*?title=\"(.*?)\"'
                               
                               '.*?<p class="star">(.*?)</p>'
                               '.*?"releasetime">(.*?)</p>'
                               '.*?"integer">(.*?)</i>.*?<i class="fraction">(.*?)</i>',
                               str(html),re.S)#不能有匹配的内容
    for items in content_lists:
        item_5 = items[5]+items[6]
        item ={'rank' : items[0],
                'title':items[1],
                'img' : items[2],
                'actor' :re.sub('\s|n|\\\\','',  items[3] ),#两个反斜杠去掉两个反斜杠
                'time' : items[4],
                'core' : item_5}
        content_list.append(item)
    return content_list
#一旦寻找顺序出错就会缺位，如img与title位置互换即title在前。


#print(content_one())
#主演本身有空格
#core出现空‘’而没有评分是因为没有后界定符

#def workhouse(item):
   # with open('movie_100.txt', 'a')as f:#用w不行，因为清楚之前，但是a可以后加,一个页面一个页面写入不能w
       #f.write(json.dumps(item, ensure_ascii=False))
       # f.close()
#opeenpyxl 存储
def excel_store(items):
    movie_100 =[]
    wb = openpyxl.Workbook()
    ws = wb.active
    #for item in items:#产生字典
        #print(item['rank'])字典没有【1】而是[‘键’]
        #for values in item.values():

            #print(values)#遍历出来的已经是字符串，而不是字典
    for i in range(10):
        for m in range(6):
                #ws.cell(row=0, column=0, value=10)必须是1开头
            ws.cell(row=i+1, column=m+1, value=str(items))
    wb.save('测试成功.xls')


#主函数
def main(number):
    url = 'https://maoyan.com/board/4?offset=' + str(number)
    html = page(url)
    items=content_one(html)#若不要workhouse导致只有最后10部电影，因为没有全部写入
    excel_store(items)




if __name__ =='__main__':
    for i in tqdm(range(10)):
        main(number=i*10)
        time.sleep(1)

#注意：被给定的形参与下一个函数传递的形参需要一致才能够调用













#内容匹配
'''
class Message_list():
    def __init__(self,movie_content):
        self.movie_content = []

    def rank_list(self):
       rank_list= re.findall('<dd>.*?board-index.*?>(.*?)</i>', str(page()),re.S)
       return rank_list

    #图片分析，第一个图片是猫眼电影，第二个才是
    def img_list(self):
        img_list = re.findall('<img data-src=\"(.*?)\"', str(page()),re.S)
        return img_list

    def title_list(self):
        title_list = re.findall('<p class="name">.*?title=\"(.*?)\"', str(page()),re.S)
        return title_list

    def role(self):
        role_list =re.findall('<p class="star">(.*?)</p>', str(page()),re.S)
        return role_list

    def release_date(self):
        release_date = re.findall('<p class="releasetime">(.*?)</p>', str(page()),re.S)
        return release_date

    def core(self):
        core_1 =  re.findall('<i class="integer">(.*?)</i>', str(page()),re.S)
        core_2 = re.findall('<i class="fraction">(.*?)</i></p>', str(page()),re.S)
        core = []
        for core_1_1 in core_1:
            for core_2_2 in core_2:
                cores = core_1_1 + core_2_2
            core.append(cores)
        return core


list_movie = Message_list(page())
print(list_movie.rank_list(),list_movie.img_list(), list_movie.title_list(),list_movie.role(),list_movie.release_date(),
    list_movie.core())
'''


#打印内容
#出现错误是因为之前已经list化了,本身findall就会列表化,寻找的是group里面的对象
